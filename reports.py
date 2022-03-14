import psycopg2
from config import config
import json
import openpyxl, pandas as pd, os
import tqdm

path = "./report_files/"
queries_file = "db_queries.json"
output_file_name = ""

if not os.path.isdir(path):
    os.mkdir(path)

def csv_to_excel(file_):
    df = pd.read_csv(file_)
    file_name = file_.replace(".csv", ".xlsx")
    df.to_excel(file_name)
    os.remove(file_)

def get_report(output):

    conn = None
    with open(queries_file, "r") as f:
        queries = json.load(f)

    wb = openpyxl.Workbook()
    try:

        params  = config()
        # print(params)
        conn = psycopg2.connect(**params)

        cur = conn.cursor()
        queries2 = tqdm.tqdm(queries)
        for query in queries2:
            #Execute Queries            
            statement = queries[query]        
            outputquery = "COPY ({0}) TO STDOUT WITH CSV HEADER".format(statement)

            

            #Write query results to csv
            with open(f"{path + query}.csv", "w") as file_:
                cur.copy_expert(outputquery, file_)
            
            #Convert result csv to excel            
            csv_to_excel(f"{path + query}.csv")

            #COPY RESULTS INTO GENERAL REPORT FILE
            queries2.set_description(f"Copying {query}")
            wb2 = openpyxl.load_workbook(f"{path + query}.xlsx")
            sh = wb2.active
            cs = wb.create_sheet(f"{query}")

            max_row = sh.max_row
            max_cols = sh.max_column
            for row in range(1, max_row + 1):
                for col in range(2, max_cols+1):
                    vals = sh.cell(row = row, column = col)
                    #print(vals.value)

                    cs.cell(row = row, column = col).value = vals.value
            os.remove(f"{path + query}.xlsx")

        wb.save(f"{path + output}.xlsx")
    except (Exception, psycopg2.DatabaseError) as error:
            print(error)
            wb.save(f"{path + output}.xlsx")
    finally:
        if conn is not None:
            conn.close()

get_report(output_file_name)


get_report_v2(output):
    conn = None
    with open("queries_file.json", "r") as f:
        queries = json.load(f)
        
    report = dict()
    try:
        params  = config()
        conn = psycopg2.connect(**params)

        cur = conn.cursor()
        queries2 = tqdm.tqdm(queries)
        for query in queries:
            statement = queries[query]
            cur.execute(statement)
            cols = [i[0] for i in cur.description]
            print(cols)
            results = []
            for row in cur.fetchall():
                results.append(dict(zip(cols, row)))

            report[query] = results
            
            with open(f"{output}.json", "w") as file_:
            json.dump(report, file_, indent=2, default=str)
    except (Exception, psycopg2.DatabaseError) as error:
            print(error)
            wb.save(f"{path + output}.xlsx")
    finally:
        if conn is not None:
            conn.close()

get_report_v2("query_report")

def convert_html():
    data = open('query_report.json','r')
    jsonFile = data.read()
    res = json.loads(jsonFile)
    html_data = json2html.convert(json=res)
    with open("Query_report_new.html", "w") as ft:
        ft.write(html_data)
        
convert_html()


